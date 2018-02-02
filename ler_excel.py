from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(filename="planilha.xlsx")
ws = wb._sheets[0]




"""
for row in range(1, 20):
	for col in range(1, 54):
		ws.cell(
			column=col,
			row=row,
			value="{}".format(get_column_letter(col))
		)
"""

wb.save("planilha.xlsx")
